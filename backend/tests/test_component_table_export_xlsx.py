"""Exact-value workbook builder used by Component Table download."""
import io
import sys
from pathlib import Path

import openpyxl

BACKEND = Path(__file__).resolve().parents[1]
if str(BACKEND) not in sys.path:
    sys.path.insert(0, str(BACKEND))

from export_routes import build_multi_sheet_xlsx  # noqa: E402
from conftest import auth_headers  # noqa: E402


def test_component_table_workbook_keeps_exact_numeric_values():
    data = build_multi_sheet_xlsx([
        {
            "name": "Component_Summary",
            "columns": [
                "component", "uom", "phys_target", "phys_achieved", "phys_percent",
                "fin_released", "fin_utilized", "fin_percent",
            ],
            "headers": [
                "Component", "Unit", "Phys Target", "Phys Achieved", "Phys %",
                "Funds Released(Cr)", "Funds Utilized(Cr)", "Util %",
            ],
            "rows": [{
                "component": "e-Sewa Kendras",
                "uom": "Count",
                "phys_target": 100,
                "phys_achieved": 42,
                "phys_percent": 42.0,
                "fin_released": 11.9,
                "fin_utilized": 8.5,
                "fin_percent": 71.4285714286,
            }],
        },
        {
            "name": "Financial_Exact",
            "columns": ["high_court", "component", "fund_released", "fund_utilized"],
            "headers": ["High Court", "Component", "Released (Cr)", "Utilised (Cr)"],
            "rows": [{
                "high_court": "Manipur",
                "component": "e-Sewa Kendras",
                "fund_released": 5.123456789,
                "fund_utilized": 7.987654321,
            }],
        },
        {
            "name": "Physical_Exact",
            "columns": ["high_court", "indicator", "target", "achieved"],
            "headers": ["High Court", "Indicator", "Target", "Achieved"],
            "rows": [{
                "high_court": "Manipur",
                "indicator": "eSewa Kendras",
                "target": 12,
                "achieved": 9,
            }],
        },
    ])
    assert data[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Component_Summary", "Financial_Exact", "Physical_Exact"]
    summary = wb["Component_Summary"]
    assert summary["A2"].value == "e-Sewa Kendras"
    assert summary["C2"].value == 100
    assert summary["D2"].value == 42
    fin = wb["Financial_Exact"]
    assert fin["C2"].value == 5.123456789
    phys = wb["Physical_Exact"]
    assert phys["C2"].value == 12
    assert phys["D2"].value == 9


def test_export_component_table_endpoint(client, viewer_session):
    token = viewer_session["token"]
    resp = client.get(
        "/api/export/dashboard/component-table",
        headers=auth_headers(token),
        params={"reporting_period": "2026-05"},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    assert resp.content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Component_Summary" in wb.sheetnames
    assert "Financial_Exact" in wb.sheetnames
    assert "Physical_Exact" in wb.sheetnames

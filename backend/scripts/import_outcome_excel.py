#!/usr/bin/env python3
"""Regenerate outcome_baseline in seed_data.json from the Phase-4 Excel dummy file."""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = ROOT.parent
DEFAULT_XLSX = REPO_ROOT / "eCourts Phase4 Outcome Tracker Dummy Data.xlsx"
SEED_PATH = ROOT / "seed_data.json"

sys.path.insert(0, str(ROOT))
from outcome_excel import parse_outcome_excel_rows  # noqa: E402


def load_excel_rows(path: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        raise SystemExit(f"Excel file not found: {xlsx}")

    rows = load_excel_rows(xlsx)
    outcome_rows = parse_outcome_excel_rows(rows)
    print(f"Parsed {len(outcome_rows)} outcome rows from {xlsx.name}")

    with open(SEED_PATH) as f:
        seed = json.load(f)
    seed["outcome_baseline"] = outcome_rows

    with open(SEED_PATH, "w") as f:
        json.dump(seed, f, indent=1, ensure_ascii=False)
        f.write("\n")
    print(f"Updated {SEED_PATH}")


if __name__ == "__main__":
    main()

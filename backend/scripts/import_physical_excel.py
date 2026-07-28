#!/usr/bin/env python3
"""Physical Tracker ETL CLI — wide DoJ Excel → PMIS.

Pipeline stages (reusable):
  1. EXTRACT  — read sheet Physical_Tracker (2-row header)
  2. TRANSFORM — HC aliases, Digitization pages→Cr, eSewa DPR/CPC → long rows
  3. LOAD     — sinks:
       --write-bulk-xlsx   Admin bulk template
       --update-seed       merge physical_baseline in seed_data.json
       --load-api          POST /api/physical/bulk (dry-run → confirm)

Example:
  python scripts/import_physical_excel.py --dry-run
  python scripts/import_physical_excel.py --write-bulk-xlsx --update-seed --load-api \\
      --api-url https://ecourt.demoapi.agrayianailabs.com \\
      --admin-email admin@pmis.gov.in --admin-password '...' \\
      --period 2025-09
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = ROOT.parent
DEFAULT_XLSX = REPO_ROOT / "Physicial_Tracker_Data_for_Achieved_till_Sep-2025.xlsx"
DEFAULT_OUT = REPO_ROOT / "etl_output" / "physical_achieved_till_sep_2025_bulk.xlsx"
SEED_PATH = ROOT / "seed_data.json"
DEFAULT_SHEET = "Physical_Tracker"
DEFAULT_PERIOD = "2025-09"

sys.path.insert(0, str(ROOT))
from physical_excel import (  # noqa: E402
    merge_into_physical_baseline,
    records_to_bulk_rows,
    transform_physical_achieved_sep2025_rows,
)


def extract_sheet_rows(path: Path, sheet_name: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def write_bulk_xlsx(records: list[dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Physical_Bulk"
    for row in records_to_bulk_rows(records):
        ws.append(row)
    wb.save(out_path)


def update_seed(records: list[dict]) -> dict:
    with open(SEED_PATH) as f:
        seed = json.load(f)
    baseline = seed.get("physical_baseline") or []
    merged, stats = merge_into_physical_baseline(baseline, records)
    seed["physical_baseline"] = merged
    with open(SEED_PATH, "w") as f:
        json.dump(seed, f, indent=1, ensure_ascii=False)
        f.write("\n")
    return stats


def load_via_api(
    bulk_path: Path,
    *,
    api_url: str,
    email: str,
    password: str,
    period: str,
    dry_run_only: bool,
) -> dict:
    try:
        import httpx
    except ImportError:
        try:
            import requests as httpx  # type: ignore
        except ImportError as exc:
            raise SystemExit("httpx or requests is required for --load-api") from exc

    base = api_url.rstrip("/")
    session_factory = getattr(httpx, "Client", None) or getattr(httpx, "Session")
    with session_factory() as client:
        login_kwargs = {"json": {"email": email, "password": password}}
        timeout = 120.0 if session_factory.__name__ == "Client" else 120
        login = client.post(f"{base}/api/auth/login", timeout=timeout, **login_kwargs)
        if login.status_code != 200:
            raise SystemExit(f"Login failed ({login.status_code}): {login.text[:300]}")

        raw = bulk_path.read_bytes()
        files = {
            "file": (
                bulk_path.name,
                raw,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        url = f"{base}/api/physical/bulk"
        preview = client.post(
            url, params={"reporting_period": period, "dry_run": "true"}, files=files, timeout=timeout
        )
        if preview.status_code != 200:
            raise SystemExit(f"Dry-run failed ({preview.status_code}): {preview.text[:500]}")
        preview_body = preview.json()
        result = {"dry_run": preview_body}
        if dry_run_only:
            return result
        token = preview_body.get("preview_token")
        if token:
            commit = client.post(
                url,
                params={"reporting_period": period, "dry_run": "false", "preview_token": token},
                timeout=timeout,
            )
        else:
            commit = client.post(
                url,
                params={"reporting_period": period, "dry_run": "false"},
                files=files,
                timeout=timeout,
            )
        if commit.status_code != 200:
            raise SystemExit(f"Commit failed ({commit.status_code}): {commit.text[:500]}")
        result["commit"] = commit.json()
        return result


def main() -> None:
    p = argparse.ArgumentParser(description="ETL: Physical Achieved till Sep 2025 → PMIS")
    p.add_argument("--source", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.add_argument("--period", default=DEFAULT_PERIOD, help="Target reporting_period YYYY-MM")
    p.add_argument("--write-bulk-xlsx", action="store_true")
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    p.add_argument("--update-seed", action="store_true")
    p.add_argument("--load-api", action="store_true")
    p.add_argument("--api-url", default="http://localhost:8001")
    p.add_argument("--admin-email", default="admin@pmis.gov.in")
    p.add_argument("--admin-password", default="")
    p.add_argument("--api-dry-run-only", action="store_true")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--exclude-zero", action="store_true")
    args = p.parse_args()

    if not args.source.exists():
        raise SystemExit(f"Source Excel not found: {args.source}")

    print("=== ETL stage 1: EXTRACT ===")
    print(f"  source: {args.source}")
    print(f"  sheet:  {args.sheet}")
    rows = extract_sheet_rows(args.source, args.sheet)
    print(f"  rows:   {len(rows)} (incl. headers)")

    print("=== ETL stage 2: TRANSFORM ===")
    result = transform_physical_achieved_sep2025_rows(
        rows, include_zero_achieved=not args.exclude_zero
    )
    for k, v in result["stats"].items():
        print(f"  {k}: {v}")
    if result["issues"]:
        for issue in result["issues"][:20]:
            print(f"  ISSUE: {issue}")
        raise SystemExit("Transform failed: unmapped High Court(s).")

    records = result["records"]
    print(f"  period target: {args.period}")
    print(f"  sample: {records[0] if records else None}")

    if args.dry_run and not (args.write_bulk_xlsx or args.update_seed or args.load_api):
        print("=== DRY RUN complete (no load) ===")
        return

    print("=== ETL stage 3: LOAD ===")
    if args.write_bulk_xlsx or args.load_api:
        write_bulk_xlsx(records, args.out)
        print(f"  wrote bulk xlsx: {args.out} ({len(records)} data rows)")

    if args.update_seed:
        seed_stats = update_seed(records)
        print(f"  seed merge: {seed_stats} → {SEED_PATH}")

    if args.load_api:
        if not args.admin_password:
            raise SystemExit("--admin-password is required with --load-api")
        if not args.out.exists():
            write_bulk_xlsx(records, args.out)
        api_result = load_via_api(
            args.out,
            api_url=args.api_url,
            email=args.admin_email,
            password=args.admin_password,
            period=args.period,
            dry_run_only=args.api_dry_run_only,
        )
        dry = api_result.get("dry_run") or {}
        print(
            "  API dry-run:",
            {k: dry.get(k) for k in ("inserted", "updated", "skipped", "reporting_period", "dry_run") if k in dry},
        )
        if "commit" in api_result:
            c = api_result["commit"]
            print(
                "  API commit:",
                {k: c.get(k) for k in ("inserted", "updated", "skipped", "reporting_period") if k in c},
            )
        else:
            print("  API commit: skipped (--api-dry-run-only)")

    print("=== ETL complete ===")


if __name__ == "__main__":
    main()

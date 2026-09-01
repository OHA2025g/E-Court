#!/usr/bin/env python3
"""Consolidated Physical+Financial Tracker ETL CLI.

Source: 2-sheet Excel with Sr. No. first
  • Physical Tracker — Court | Component | Description | Target | Achieved | Remarks
  • Financial Tracker — Court | Component | Funds Released/Utilised (₹) | Remarks

Pipeline:
  1. EXTRACT  — openpyxl both sheets
  2. TRANSFORM — HC/component/indicator aliases, messy numbers, ₹→Cr, Rooms+Complex sum
  3. LOAD     — bulk xlsx and/or Admin /api/{physical|financial}/bulk

Example:
  python scripts/import_consolidated_tracker_excel.py --dry-run
  python scripts/import_consolidated_tracker_excel.py --write-bulk-xlsx --load-api \\
      --api-url http://localhost:8003 \\
      --admin-email admin@pmis.gov.in --admin-password '...' \\
      --period 2026-06
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = ROOT.parent
DEFAULT_XLSX = REPO_ROOT / "Physical_Financial_Tracker_Consolidated_2_Sheets_SrNo_First.xlsx"
DEFAULT_PHYS_OUT = REPO_ROOT / "etl_output" / "physical_consolidated_bulk.xlsx"
DEFAULT_FIN_OUT = REPO_ROOT / "etl_output" / "financial_consolidated_bulk.xlsx"
DEFAULT_PERIOD = "2026-06"

sys.path.insert(0, str(ROOT))
from consolidated_tracker_excel import (  # noqa: E402
    financial_records_to_bulk_rows,
    physical_records_to_bulk_rows,
    transform_consolidated_financial_rows,
    transform_consolidated_physical_rows,
)


def extract_sheet_rows(path: Path, sheet_name: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def write_xlsx(headers_and_rows: list[list], out_path: Path, sheet_title: str) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    for row in headers_and_rows:
        ws.append(row)
    wb.save(out_path)


def load_via_api(
    bulk_path: Path,
    *,
    tracker: str,
    api_url: str,
    email: str,
    password: str,
    period: str,
    dry_run_only: bool,
) -> dict:
    try:
        import httpx
    except ImportError:
        try:
            import requests as httpx  # type: ignore
        except ImportError as exc:
            raise SystemExit("httpx or requests is required for --load-api") from exc

    base = api_url.rstrip("/")
    session_factory = getattr(httpx, "Client", None) or getattr(httpx, "Session")
    with session_factory() as client:
        timeout = 180.0 if session_factory.__name__ == "Client" else 180
        login = client.post(
            f"{base}/api/auth/login",
            json={"email": email, "password": password},
            timeout=timeout,
        )
        if login.status_code != 200:
            raise SystemExit(f"Login failed ({login.status_code}): {login.text[:300]}")

        raw = bulk_path.read_bytes()
        files = {
            "file": (
                bulk_path.name,
                raw,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        url = f"{base}/api/{tracker}/bulk"
        preview = client.post(
            url,
            params={"reporting_period": period, "dry_run": "true"},
            files=files,
            timeout=timeout,
        )
        if preview.status_code != 200:
            raise SystemExit(f"{tracker} dry-run failed ({preview.status_code}): {preview.text[:800]}")
        preview_body = preview.json()
        result = {"dry_run": preview_body}
        if dry_run_only:
            return result
        token = preview_body.get("preview_token")
        if token:
            commit = client.post(
                url,
                params={"reporting_period": period, "dry_run": "false", "preview_token": token},
                timeout=timeout,
            )
        else:
            commit = client.post(
                url,
                params={"reporting_period": period, "dry_run": "false"},
                files=files,
                timeout=timeout,
            )
        if commit.status_code != 200:
            raise SystemExit(f"{tracker} commit failed ({commit.status_code}): {commit.text[:800]}")
        result["commit"] = commit.json()
        return result


def _print_summary(label: str, body: dict) -> None:
    summary = (body or {}).get("summary") or {}
    print(
        f"  {label}: inserted={summary.get('inserted') or body.get('inserted')} "
        f"updated={summary.get('updated') or body.get('updated')} "
        f"skipped={summary.get('skipped') or body.get('skipped')} "
        f"valid={summary.get('valid')} invalid={summary.get('invalid')}"
    )
    errs = (body or {}).get("errors") or []
    for e in errs[:12]:
        print(f"    ERR: {e}")


def main() -> None:
    p = argparse.ArgumentParser(description="ETL: Consolidated 2-sheet tracker Excel → PMIS")
    p.add_argument("--source", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--period", default=DEFAULT_PERIOD, help="Target reporting_period YYYY-MM")
    p.add_argument("--physical-sheet", default="Physical Tracker")
    p.add_argument("--financial-sheet", default="Financial Tracker")
    p.add_argument("--tracker", choices=("both", "physical", "financial"), default="both")
    p.add_argument("--write-bulk-xlsx", action="store_true")
    p.add_argument("--phys-out", type=Path, default=DEFAULT_PHYS_OUT)
    p.add_argument("--fin-out", type=Path, default=DEFAULT_FIN_OUT)
    p.add_argument("--load-api", action="store_true")
    p.add_argument("--api-url", default="http://localhost:8003")
    p.add_argument("--admin-email", default="admin@pmis.gov.in")
    p.add_argument("--admin-password", default="")
    p.add_argument("--api-dry-run-only", action="store_true")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    if not args.source.exists():
        raise SystemExit(f"Source Excel not found: {args.source}")

    do_phys = args.tracker in ("both", "physical")
    do_fin = args.tracker in ("both", "financial")

    phys_records = []
    fin_records = []

    if do_phys:
        print("=== PHYSICAL EXTRACT ===")
        print(f"  source: {args.source}")
        print(f"  sheet:  {args.physical_sheet}")
        phys_rows = extract_sheet_rows(args.source, args.physical_sheet)
        print(f"  rows:   {len(phys_rows)} (incl. header)")
        print("=== PHYSICAL TRANSFORM ===")
        phys_result = transform_consolidated_physical_rows(phys_rows)
        for k, v in phys_result["stats"].items():
            print(f"  {k}: {v}")
        if phys_result["issues"]:
            print(f"  soft issues ({len(phys_result['issues'])}):")
            for iss in phys_result["issues"][:15]:
                print(f"    {iss}")
        if phys_result["stats"].get("unknown_high_courts"):
            raise SystemExit("Transform failed: unmapped High Court(s).")
        if phys_result["stats"].get("unknown_components"):
            raise SystemExit("Transform failed: unmapped Component(s).")
        phys_records = phys_result["records"]
        print(f"  sample: {phys_records[0] if phys_records else None}")

    if do_fin:
        print("=== FINANCIAL EXTRACT ===")
        print(f"  source: {args.source}")
        print(f"  sheet:  {args.financial_sheet}")
        fin_rows = extract_sheet_rows(args.source, args.financial_sheet)
        print(f"  rows:   {len(fin_rows)} (incl. header)")
        print("=== FINANCIAL TRANSFORM ===")
        fin_result = transform_consolidated_financial_rows(fin_rows)
        for k, v in fin_result["stats"].items():
            print(f"  {k}: {v}")
        if fin_result["issues"]:
            print(f"  soft issues ({len(fin_result['issues'])}):")
            for iss in fin_result["issues"][:15]:
                print(f"    {iss}")
        if fin_result["stats"].get("unknown_high_courts"):
            raise SystemExit("Transform failed: unmapped High Court(s).")
        if fin_result["stats"].get("unknown_components"):
            raise SystemExit("Transform failed: unmapped Component(s).")
        fin_records = fin_result["records"]
        print(f"  sample: {fin_records[0] if fin_records else None}")

    print(f"  period target: {args.period}")

    if args.dry_run and not (args.write_bulk_xlsx or args.load_api):
        print("=== DRY RUN complete (no load) ===")
        return

    print("=== LOAD ===")
    if do_phys and (args.write_bulk_xlsx or args.load_api):
        write_xlsx(physical_records_to_bulk_rows(phys_records), args.phys_out, "Physical_Bulk")
        print(f"  wrote {args.phys_out} ({len(phys_records)} rows)")
    if do_fin and (args.write_bulk_xlsx or args.load_api):
        write_xlsx(financial_records_to_bulk_rows(fin_records, amounts_as_rupees=True), args.fin_out, "Financial_Bulk")
        print(f"  wrote {args.fin_out} ({len(fin_records)} rows)")

    if args.load_api:
        if not args.admin_password:
            raise SystemExit("--admin-password is required with --load-api")
        if do_phys:
            print("=== API physical bulk ===")
            body = load_via_api(
                args.phys_out,
                tracker="physical",
                api_url=args.api_url,
                email=args.admin_email,
                password=args.admin_password,
                period=args.period,
                dry_run_only=args.api_dry_run_only,
            )
            _print_summary("dry_run", body.get("dry_run") or {})
            if body.get("commit"):
                _print_summary("commit", body["commit"])
        if do_fin:
            print("=== API financial bulk ===")
            body = load_via_api(
                args.fin_out,
                tracker="financial",
                api_url=args.api_url,
                email=args.admin_email,
                password=args.admin_password,
                period=args.period,
                dry_run_only=args.api_dry_run_only,
            )
            _print_summary("dry_run", body.get("dry_run") or {})
            if body.get("commit"):
                _print_summary("commit", body["commit"])

    print("=== DONE ===")


if __name__ == "__main__":
    main()

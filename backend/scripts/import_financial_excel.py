#!/usr/bin/env python3
"""Financial Tracker ETL CLI — wide DoJ Excel → PMIS (Released or Utilised).

Pipeline stages (reusable):
  1. EXTRACT  — read wide sheet from source .xlsx
  2. TRANSFORM — HC/component alias map, unpivot to long rows, ₹ → ₹ Cr
  3. LOAD     — one or more sinks:
       --write-bulk-xlsx   long-format file for Admin Bulk Upload UI / API
       --update-seed       merge into seed_data.json baseline
       --load-api          POST /api/financial/bulk (dry-run then confirm)

Modes:
  --mode released   (default) Funds Released → fund_released
  --mode utilised   Funds Utilised → fund_utilized

Examples:
  python scripts/import_financial_excel.py --dry-run

  python scripts/import_financial_excel.py --mode utilised \\
      --source ../Financial_Tracker_Data_for_Utilised_Fund_2023-2024.xlsx \\
      --sheet Funds_Utilised --write-bulk-xlsx --update-seed --load-api \\
      --api-url https://ecourt.demoapi.agrayianailabs.com \\
      --admin-email admin@pmis.gov.in --admin-password '...' \\
      --period 2026-05
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = ROOT.parent
DEFAULT_RELEASED_XLSX = REPO_ROOT / "Financial_Tracker_Data_for_Released_Fund_2024-2027.xlsx"
DEFAULT_UTILISED_XLSX = REPO_ROOT / "Financial_Tracker_Data_for_Utilised_Fund_2023-2024.xlsx"
SEED_PATH = ROOT / "seed_data.json"
DEFAULT_PERIOD = "2026-05"

sys.path.insert(0, str(ROOT))
from financial_excel import (  # noqa: E402
    FUND_FIELD_RELEASED,
    FUND_FIELD_UTILIZED,
    merge_released_into_seed_baseline,
    merge_utilised_into_seed_baseline,
    records_to_bulk_rows,
    transform_funds_released_rows,
    transform_funds_utilised_rows,
)


def extract_sheet_rows(path: Path, sheet_name: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def write_bulk_xlsx(records: list[dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial_Bulk"
    for row in records_to_bulk_rows(records):
        ws.append(row)
    wb.save(out_path)


def update_seed(records: list[dict], fund_field: str) -> dict:
    with open(SEED_PATH) as f:
        seed = json.load(f)
    baseline = seed.get("financial_baseline") or []
    if fund_field == FUND_FIELD_UTILIZED:
        merged, stats = merge_utilised_into_seed_baseline(baseline, records)
    else:
        merged, stats = merge_released_into_seed_baseline(baseline, records)
    seed["financial_baseline"] = merged
    with open(SEED_PATH, "w") as f:
        json.dump(seed, f, indent=1, ensure_ascii=False)
        f.write("\n")
    return stats


def enrich_counterpart_from_seed(records: list[dict], primary_field: str) -> int:
    """Fill the blank counterpart fund column from seed so older bulk APIs do not wipe it."""
    counterpart = (
        FUND_FIELD_RELEASED if primary_field == FUND_FIELD_UTILIZED else FUND_FIELD_UTILIZED
    )
    if not SEED_PATH.exists():
        return 0
    with open(SEED_PATH) as f:
        seed = json.load(f)
    by_key = {
        (r["high_court"], r["component"]): r
        for r in (seed.get("financial_baseline") or [])
    }
    filled = 0
    for rec in records:
        if rec.get(counterpart) is not None:
            continue
        existing = by_key.get((rec["high_court"], rec["component"]))
        if existing and existing.get(counterpart) is not None:
            rec[counterpart] = existing[counterpart]
            filled += 1
    return filled


def load_via_api(
    bulk_path: Path,
    *,
    api_url: str,
    email: str,
    password: str,
    period: str,
    dry_run_only: bool,
) -> dict:
    try:
        import httpx
    except ImportError:
        try:
            import requests as httpx  # type: ignore
        except ImportError as exc:
            raise SystemExit("httpx or requests is required for --load-api") from exc

    base = api_url.rstrip("/")
    session_factory = getattr(httpx, "Client", None) or getattr(httpx, "Session")
    with session_factory() as client:
        login_kwargs = {"json": {"email": email, "password": password}}
        if session_factory.__name__ == "Client":
            login = client.post(f"{base}/api/auth/login", timeout=60.0, **login_kwargs)
        else:
            login = client.post(f"{base}/api/auth/login", timeout=60, **login_kwargs)
        if login.status_code != 200:
            raise SystemExit(f"Login failed ({login.status_code}): {login.text[:300]}")

        raw = bulk_path.read_bytes()
        files = {
            "file": (
                bulk_path.name,
                raw,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        preview_url = f"{base}/api/financial/bulk"
        params = {"reporting_period": period, "dry_run": "true"}
        if session_factory.__name__ == "Client":
            preview = client.post(preview_url, params=params, files=files, timeout=120.0)
        else:
            preview = client.post(preview_url, params=params, files=files, timeout=120)
        if preview.status_code != 200:
            raise SystemExit(f"Dry-run failed ({preview.status_code}): {preview.text[:500]}")
        preview_body = preview.json()
        token = preview_body.get("preview_token")
        result = {"dry_run": preview_body}
        if dry_run_only:
            return result
        if not token:
            commit_params = {"reporting_period": period, "dry_run": "false"}
            if session_factory.__name__ == "Client":
                commit = client.post(preview_url, params=commit_params, files=files, timeout=120.0)
            else:
                commit = client.post(preview_url, params=commit_params, files=files, timeout=120)
        else:
            commit_params = {"reporting_period": period, "dry_run": "false", "preview_token": token}
            if session_factory.__name__ == "Client":
                commit = client.post(preview_url, params=commit_params, timeout=120.0)
            else:
                commit = client.post(preview_url, params=commit_params, timeout=120)
        if commit.status_code != 200:
            raise SystemExit(f"Commit failed ({commit.status_code}): {commit.text[:500]}")
        result["commit"] = commit.json()
        return result


def main() -> None:
    p = argparse.ArgumentParser(description="ETL: DoJ Financial Excel → PMIS Financial Tracker")
    p.add_argument(
        "--mode",
        choices=("released", "utilised", "utilized"),
        default="released",
        help="Which fund field to load (default: released)",
    )
    p.add_argument("--source", type=Path, default=None, help="Wide DoJ Excel path")
    p.add_argument("--sheet", default=None, help="Sheet name")
    p.add_argument("--period", default=DEFAULT_PERIOD, help="Target reporting_period YYYY-MM")
    p.add_argument("--write-bulk-xlsx", action="store_true", help="Write long-format bulk Excel")
    p.add_argument("--out", type=Path, default=None, help="Bulk xlsx output path")
    p.add_argument("--update-seed", action="store_true", help="Merge into seed_data.json")
    p.add_argument("--load-api", action="store_true", help="Upload via Admin bulk API")
    p.add_argument("--api-url", default="http://localhost:8001")
    p.add_argument("--admin-email", default="admin@pmis.gov.in")
    p.add_argument("--admin-password", default="")
    p.add_argument("--api-dry-run-only", action="store_true", help="Stop after API dry-run")
    p.add_argument("--dry-run", action="store_true", help="Extract+transform only (print stats)")
    p.add_argument("--exclude-zero", action="store_true", help="Skip zero amounts")
    args = p.parse_args()

    mode = "utilised" if args.mode in ("utilised", "utilized") else "released"
    fund_field = FUND_FIELD_UTILIZED if mode == "utilised" else FUND_FIELD_RELEASED

    if args.source is None:
        args.source = DEFAULT_UTILISED_XLSX if mode == "utilised" else DEFAULT_RELEASED_XLSX
    if args.sheet is None:
        args.sheet = "Funds_Utilised" if mode == "utilised" else "Funds_Released"
    if args.out is None:
        name = (
            "financial_funds_utilised_2023_2024_bulk.xlsx"
            if mode == "utilised"
            else "financial_funds_released_2024_2027_bulk.xlsx"
        )
        args.out = REPO_ROOT / "etl_output" / name

    if not args.source.exists():
        raise SystemExit(f"Source Excel not found: {args.source}")

    print("=== ETL stage 1: EXTRACT ===")
    print(f"  mode:   {mode} → {fund_field}")
    print(f"  source: {args.source}")
    print(f"  sheet:  {args.sheet}")
    rows = extract_sheet_rows(args.source, args.sheet)
    print(f"  rows:   {len(rows)} (incl. header)")

    print("=== ETL stage 2: TRANSFORM ===")
    if mode == "utilised":
        result = transform_funds_utilised_rows(rows, include_zero=not args.exclude_zero)
    else:
        result = transform_funds_released_rows(rows, include_zero=not args.exclude_zero)
    stats = result["stats"]
    for k, v in stats.items():
        print(f"  {k}: {v}")
    if result["issues"]:
        print(f"  ISSUES ({len(result['issues'])}):")
        for issue in result["issues"][:20]:
            print(f"    - {issue}")
        raise SystemExit("Transform failed: unmapped High Court(s). Fix aliases and retry.")

    records = result["records"]
    print(f"  period target: {args.period}")
    print(f"  sample: {records[0] if records else None}")

    if args.dry_run and not (args.write_bulk_xlsx or args.update_seed or args.load_api):
        print("=== DRY RUN complete (no load) ===")
        return

    print("=== ETL stage 3: LOAD ===")
    filled = enrich_counterpart_from_seed(records, fund_field)
    if filled:
        counterpart = "fund_released" if fund_field == FUND_FIELD_UTILIZED else "fund_utilized"
        print(f"  enriched {counterpart} from seed: {filled} rows")

    if args.write_bulk_xlsx or args.load_api:
        write_bulk_xlsx(records, args.out)
        print(f"  wrote bulk xlsx: {args.out} ({len(records)} data rows)")

    if args.update_seed:
        seed_stats = update_seed(records, fund_field)
        print(f"  seed merge: {seed_stats} → {SEED_PATH}")

    if args.load_api:
        if not args.admin_password:
            raise SystemExit("--admin-password is required with --load-api")
        if not args.out.exists():
            write_bulk_xlsx(records, args.out)
        api_result = load_via_api(
            args.out,
            api_url=args.api_url,
            email=args.admin_email,
            password=args.admin_password,
            period=args.period,
            dry_run_only=args.api_dry_run_only,
        )
        dry = api_result.get("dry_run") or {}
        print(
            "  API dry-run:",
            {k: dry.get(k) for k in ("inserted", "updated", "skipped", "reporting_period", "dry_run") if k in dry},
        )
        if "commit" in api_result:
            c = api_result["commit"]
            print(
                "  API commit:",
                {k: c.get(k) for k in ("inserted", "updated", "skipped", "reporting_period") if k in c},
            )
        else:
            print("  API commit: skipped (--api-dry-run-only)")

    print("=== ETL complete ===")


if __name__ == "__main__":
    main()

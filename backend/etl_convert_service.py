"""Convert wide DoJ / Phase-4 Excel into PMIS long-format with 1:1 mapping preview."""
from __future__ import annotations

import io
from typing import Any, Optional

import openpyxl
from openpyxl import Workbook

from consolidated_tracker_excel import (
    financial_records_to_bulk_rows as consol_fin_bulk_rows,
    is_consolidated_financial_header,
    is_consolidated_physical_header,
    physical_records_to_bulk_rows as consol_phys_bulk_rows,
    transform_consolidated_financial_rows,
    transform_consolidated_physical_rows,
)
from financial_excel import (
    BULK_HEADERS as FIN_HEADERS,
    COMPONENT_ALIASES,
    HC_ALIASES,
    map_component,
    records_to_bulk_rows as fin_records_to_bulk_rows,
    transform_funds_released_rows,
    transform_funds_utilised_rows,
)
from outcome_excel import map_outcome_subject, parse_outcome_excel_rows
from physical_excel import (
    BULK_HEADERS as PHYS_HEADERS,
    COMPONENT_DIGITISATION,
    COMPONENT_ESEWA,
    INDICATOR_ESEWA_IN_COMPLEXES,
    INDICATOR_PAGES_DIGITIZED,
    records_to_bulk_rows as phys_records_to_bulk_rows,
    transform_physical_achieved_sep2025_rows,
)

OUTCOME_BULK_HEADERS = [
    "High Court",
    "Component",
    "Sub-Component",
    "Subject",
    "KPI ID",
    "Granularity",
    "District",
    "Value",
    "Baseline",
    "Remarks",
]

MAX_ROW_MAPPINGS = 250


def _norm(s: Any) -> str:
    return " ".join(str(s or "").strip().lower().split())


def load_workbook_sheets(raw: bytes) -> dict[str, list[tuple]]:
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    sheets: dict[str, list[tuple]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        sheets[name] = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return sheets


def pick_sheet(sheets: dict[str, list[tuple]], preferred: Optional[str] = None) -> tuple[str, list[tuple]]:
    if preferred and preferred in sheets:
        return preferred, sheets[preferred]
    for name in (
        "Physical Tracker",
        "Financial Tracker",
        "Physical_Tracker",
        "Funds_Released",
        "Funds_Utilised",
        "Funds_Utilized",
        "Sheet1",
    ):
        if name in sheets:
            return name, sheets[name]
    first = next(iter(sheets.items()))
    return first[0], first[1]


def header_cells(row: tuple | list) -> list[str]:
    return [str(c or "").strip() for c in row]


def headers_lower(row: tuple | list) -> list[str]:
    return [_norm(c) for c in row]


def is_long_physical(header: list[str]) -> bool:
    h = set(header)
    return "high court" in h and "component" in h and (
        "sub-component" in h or "sub component" in h or "indicator" in h
    ) and "achieved" in h


def is_long_financial(header: list[str]) -> bool:
    h = set(header)
    return "high court" in h and "component" in h and (
        "fund released" in h or "fund utilized" in h or "fund utilised" in h
    )


def is_long_outcome(header: list[str]) -> bool:
    h = set(header)
    return "high court" in h and "subject" in h and ("kpi id" in h or "kpid" in h) and "value" in h


def is_phase4_outcome(header: list[str]) -> bool:
    joined = " ".join(header)
    return ("kpid" in header or "kpi id" in header) and (
        "components" in header or "outcome" in joined
    )


def is_wide_physical(rows: list[tuple]) -> bool:
    if len(rows) < 3:
        return False
    r0 = " ".join(headers_lower(rows[0]))
    r1 = " ".join(headers_lower(rows[1]))
    return ("high court" in r0 or "high courts" in r0 or "high court" in r1) and (
        "digit" in r0 or "digit" in r1 or "e-sewa" in r0 or "esewa" in r0 or "e sewa" in r1
    )


def write_bulk_xlsx(headers: list[str], data_rows: list[list[Any]], sheet_name: str = "Bulk") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _source_snapshot(row: tuple | list, labels: list[str], limit: int = 8) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for i, label in enumerate(labels[:limit]):
        if not label:
            continue
        val = row[i] if i < len(row) else None
        if val is None or str(val).strip() == "":
            continue
        out[label] = val
    return out


def convert_physical(raw: bytes, *, sheet: Optional[str] = None) -> dict[str, Any]:
    sheets = load_workbook_sheets(raw)
    preferred = sheet
    if not preferred:
        if "Physical Tracker" in sheets:
            preferred = "Physical Tracker"
        elif "Physical_Tracker" in sheets:
            preferred = "Physical_Tracker"
    sheet_name, rows = pick_sheet(sheets, preferred)
    if not rows:
        raise ValueError("Empty sheet")

    header0 = headers_lower(rows[0])

    if is_long_physical(header0):
        mappings = [
            {"source": "High Court", "target": "High Court", "transform": "identity"},
            {"source": "Component", "target": "Component", "transform": "identity"},
            {"source": "Sub-Component / Indicator", "target": "Sub-Component", "transform": "identity"},
            {"source": "District", "target": "District", "transform": "identity"},
            {"source": "Target", "target": "Target", "transform": "identity"},
            {"source": "Achieved", "target": "Achieved", "transform": "identity"},
            {"source": "Remarks", "target": "Remarks", "transform": "identity"},
        ]
        row_maps = []
        labels = header_cells(rows[0])
        for i, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            src = _source_snapshot(row, labels)
            by = {_norm(labels[j]): row[j] if j < len(row) else None for j in range(len(labels))}
            target = {
                "High Court": by.get("high court"),
                "Component": by.get("component"),
                "Sub-Component": by.get("sub-component") or by.get("sub component") or by.get("indicator"),
                "District": by.get("district") or "",
                "Target": by.get("target"),
                "Achieved": by.get("achieved"),
                "Remarks": by.get("remarks") or "",
            }
            row_maps.append({"source_row": i, "source": src, "target": target, "status": "ok"})
        return {
            "tracker": "physical",
            "format_detected": "long_bulk",
            "sheet": sheet_name,
            "column_mappings": mappings,
            "row_mappings": row_maps[:MAX_ROW_MAPPINGS],
            "row_mappings_total": len(row_maps),
            "stats": {"records": len(row_maps), "source_data_rows": len(rows) - 1},
            "issues": [],
            "bulk_bytes": raw,
            "bulk_filename": "physical_bulk.xlsx",
            "desired_headers": list(PHYS_HEADERS),
        }

    if is_consolidated_physical_header(header0):
        result = transform_consolidated_physical_rows(rows)
        records = result["records"]
        issues = list(result.get("issues") or [])
        column_mappings = [
            {"source": "Court", "target": "High Court", "transform": "HC_ALIASES (Orissa→Odisha, Gauhati…)"},
            {"source": "Component", "target": "Component", "transform": "CONSOLIDATED_COMPONENT_ALIASES"},
            {"source": "Description", "target": "Sub-Component", "transform": "DESCRIPTION_INDICATOR_ALIASES"},
            {"source": "Target / Achieved", "target": "Target / Achieved (e-Sewa → DPR/CPC)", "transform": "parse_messy_quantity; pages→Cr"},
            {"source": "Physical Tracker Remarks", "target": "Remarks", "transform": "append ETL note"},
        ]
        row_maps = []
        labels = header_cells(rows[0])
        for rec in records:
            src_row_idx = rec.get("source_row") or 0
            src_row = rows[src_row_idx - 1] if 0 < src_row_idx <= len(rows) else ()
            row_maps.append({
                "source_row": src_row_idx,
                "source": _source_snapshot(src_row, labels, limit=9),
                "target": {
                    "High Court": rec["high_court"],
                    "Component": rec["component"],
                    "Sub-Component": rec["indicator"],
                    "District": rec.get("district") or "",
                    "Target": rec.get("target"),
                    "Achieved": rec.get("achieved"),
                    "Target as per DPR": rec.get("target_dpr"),
                    "Achieved as per CPC": rec.get("achieved_cpc"),
                    "Remarks": rec.get("remarks") or "",
                },
                "status": "ok",
            })
        for iss in issues:
            row_maps.append({
                "source_row": iss.get("row"),
                "source": {"error": iss.get("error")},
                "target": {},
                "status": "error",
                "error": iss.get("error"),
            })
        bulk_rows = consol_phys_bulk_rows(records)
        bulk_bytes = write_bulk_xlsx(PHYS_HEADERS, bulk_rows[1:], "Physical_Bulk")
        return {
            "tracker": "physical",
            "format_detected": "consolidated_long_physical",
            "sheet": sheet_name,
            "column_mappings": column_mappings,
            "row_mappings": row_maps[:MAX_ROW_MAPPINGS],
            "row_mappings_total": len(row_maps),
            "stats": result.get("stats") or {},
            "issues": issues,
            "bulk_bytes": bulk_bytes,
            "bulk_filename": "physical_consolidated_converted_bulk.xlsx",
            "desired_headers": list(PHYS_HEADERS),
            "alias_hints": {"high_court_aliases": len(HC_ALIASES)},
        }

    _ = is_wide_physical(rows)  # detection hint only; transform raises if unusable
    result = transform_physical_achieved_sep2025_rows(rows)
    records = result["records"]
    issues = list(result.get("issues") or [])

    column_mappings = [
        {"source": "High Courts (col B)", "target": "High Court", "transform": "HC_ALIASES"},
        {
            "source": "Digitization Target / Achieved Pgs",
            "target": f"Component={COMPONENT_DIGITISATION} · Sub-Component={INDICATOR_PAGES_DIGITIZED} · Target/Achieved",
            "transform": "absolute pages ÷ 10,000,000 → Crore Pages",
        },
        {
            "source": "eSewa Target DPR / Achieved CPC (fallback eCommittee)",
            "target": f"Component={COMPONENT_ESEWA} · Sub-Component={INDICATOR_ESEWA_IN_COMPLEXES} · Target/Achieved",
            "transform": "Absolute Count (identity)",
        },
    ]

    src_labels = [
        "Sr.No", "High Courts", "Dig Target Pgs", "Dig Achieved Pgs", "%",
        "eSewa Target DPR", "eSewa Ach eCommittee", "%", "eSewa Target CPC", "eSewa Ach CPC", "%",
    ]

    row_maps = []
    for rec in records:
        src_row_idx = rec.get("source_row") or 0
        src_row = rows[src_row_idx - 1] if 0 < src_row_idx <= len(rows) else ()
        row_maps.append({
            "source_row": src_row_idx,
            "source": _source_snapshot(src_row, src_labels, limit=11),
            "target": {
                "High Court": rec["high_court"],
                "Component": rec["component"],
                "Sub-Component": rec["indicator"],
                "District": rec.get("district") or "",
                "Target": rec.get("target"),
                "Achieved": rec.get("achieved"),
                "Remarks": rec.get("remarks") or "",
            },
            "status": "ok",
            "measure": rec.get("measure"),
        })

    for iss in issues:
        row_maps.append({
            "source_row": iss.get("row"),
            "source": {"error": iss.get("error")},
            "target": {},
            "status": "error",
            "error": iss.get("error"),
        })

    bulk_rows = phys_records_to_bulk_rows(records)
    bulk_bytes = write_bulk_xlsx(PHYS_HEADERS, bulk_rows[1:], "Physical_Bulk")

    return {
        "tracker": "physical",
        "format_detected": "wide_doj_physical",
        "sheet": sheet_name,
        "column_mappings": column_mappings,
        "row_mappings": row_maps[:MAX_ROW_MAPPINGS],
        "row_mappings_total": len(row_maps),
        "stats": result.get("stats") or {},
        "issues": issues,
        "bulk_bytes": bulk_bytes,
        "bulk_filename": "physical_converted_bulk.xlsx",
        "desired_headers": list(PHYS_HEADERS),
        "alias_hints": {"high_court_aliases": len(HC_ALIASES)},
    }


def convert_financial(
    raw: bytes,
    *,
    mode: str = "auto",
    sheet: Optional[str] = None,
) -> dict[str, Any]:
    sheets = load_workbook_sheets(raw)
    preferred = sheet
    if not preferred:
        if "Financial Tracker" in sheets:
            preferred = "Financial Tracker"
        elif mode == "utilised":
            preferred = "Funds_Utilised" if "Funds_Utilised" in sheets else "Funds_Utilized"
        elif mode == "released":
            preferred = "Funds_Released"
    sheet_name, rows = pick_sheet(sheets, preferred)
    if not rows:
        raise ValueError("Empty sheet")

    header0 = headers_lower(rows[0])
    display_header = header_cells(rows[0])

    if is_long_financial(header0):
        mappings = [
            {"source": "High Court", "target": "High Court", "transform": "identity"},
            {"source": "Component", "target": "Component", "transform": "identity"},
            {"source": "District", "target": "District", "transform": "identity"},
            {"source": "Fund Released", "target": "Fund Released", "transform": "identity (₹ Cr)"},
            {"source": "Fund Utilized", "target": "Fund Utilized", "transform": "identity (₹ Cr)"},
            {"source": "Remarks", "target": "Remarks", "transform": "identity"},
        ]
        row_maps = []
        for i, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            by = {_norm(display_header[j]): row[j] if j < len(row) else None for j in range(len(display_header))}
            util = by.get("fund utilized")
            if util is None:
                util = by.get("fund utilised")
            target = {
                "High Court": by.get("high court"),
                "Component": by.get("component"),
                "District": by.get("district") or "",
                "Fund Released": by.get("fund released"),
                "Fund Utilized": util,
                "Remarks": by.get("remarks") or "",
            }
            row_maps.append({
                "source_row": i,
                "source": _source_snapshot(row, display_header),
                "target": target,
                "status": "ok",
            })
        return {
            "tracker": "financial",
            "format_detected": "long_bulk",
            "sheet": sheet_name,
            "mode": mode,
            "column_mappings": mappings,
            "row_mappings": row_maps[:MAX_ROW_MAPPINGS],
            "row_mappings_total": len(row_maps),
            "stats": {"records": len(row_maps)},
            "issues": [],
            "bulk_bytes": raw,
            "bulk_filename": "financial_bulk.xlsx",
            "desired_headers": list(FIN_HEADERS),
        }

    if is_consolidated_financial_header(header0):
        result = transform_consolidated_financial_rows(rows)
        records = result["records"]
        issues = list(result.get("issues") or [])
        column_mappings = [
            {"source": "Court", "target": "High Court", "transform": "HC_ALIASES"},
            {
                "source": "Component",
                "target": "Component",
                "transform": "CONSOLIDATED_COMPONENT_ALIASES; Rooms+Complex summed",
            },
            {
                "source": "Funds Released (₹)",
                "target": "Fund Released",
                "transform": "₹ absolute ÷ 10,000,000 → ₹ Cr",
            },
            {
                "source": "Funds Utilised (₹)",
                "target": "Fund Utilized",
                "transform": "₹ absolute ÷ 10,000,000 → ₹ Cr",
            },
            {
                "source": "Financial Tracker Remarks",
                "target": "Remarks",
                "transform": "append ETL note",
            },
        ]
        row_maps = []
        for rec in records:
            src_row_idx = rec.get("source_row") or 0
            src_row = rows[src_row_idx - 1] if 0 < src_row_idx <= len(rows) else ()
            row_maps.append({
                "source_row": src_row_idx,
                "source": _source_snapshot(src_row, display_header, limit=9),
                "target": {
                    "High Court": rec["high_court"],
                    "Component": rec["component"],
                    "District": "",
                    "Fund Released": rec.get("fund_released"),
                    "Fund Utilized": rec.get("fund_utilized"),
                    "Remarks": rec.get("remarks") or "",
                },
                "status": "ok",
            })
        for iss in issues:
            row_maps.append({
                "source_row": iss.get("row"),
                "source": {"error": iss.get("error")},
                "target": {},
                "status": "error",
                "error": iss.get("error"),
            })
        bulk_rows = consol_fin_bulk_rows(records, amounts_as_rupees=True)
        bulk_bytes = write_bulk_xlsx(FIN_HEADERS, bulk_rows[1:], "Financial_Bulk")
        return {
            "tracker": "financial",
            "format_detected": "consolidated_long_financial",
            "sheet": sheet_name,
            "mode": "both",
            "column_mappings": column_mappings,
            "row_mappings": row_maps[:MAX_ROW_MAPPINGS],
            "row_mappings_total": len(row_maps),
            "stats": result.get("stats") or {},
            "issues": issues,
            "bulk_bytes": bulk_bytes,
            "bulk_filename": "financial_consolidated_converted_bulk.xlsx",
            "desired_headers": list(FIN_HEADERS),
            "alias_hints": {
                "component_aliases": len(COMPONENT_ALIASES),
                "high_court_aliases": len(HC_ALIASES),
            },
        }

    resolved_mode = mode
    if mode == "auto":
        name_l = sheet_name.lower()
        if "utilis" in name_l or "utiliz" in name_l:
            resolved_mode = "utilised"
        else:
            resolved_mode = "released"

    if resolved_mode == "utilised":
        result = transform_funds_utilised_rows(rows)
        fund_label = "Fund Utilized"
        fund_key = "fund_utilized"
        unit_note = "₹ absolute ÷ 10,000,000 → ₹ Cr (utilised)"
    else:
        result = transform_funds_released_rows(rows)
        fund_label = "Fund Released"
        fund_key = "fund_released"
        unit_note = "₹ absolute ÷ 10,000,000 → ₹ Cr (released)"

    records = result["records"]
    issues = list(result.get("issues") or [])

    column_mappings = [
        {"source": "High Courts", "target": "High Court", "transform": "HC_ALIASES"},
    ]
    for raw_name in display_header:
        canon = map_component(raw_name)
        if canon:
            column_mappings.append({
                "source": str(raw_name).strip(),
                "target": f"Component={canon} · {fund_label}",
                "transform": unit_note,
            })

    # Find HC column for source snapshot
    hc_col = 0
    for idx, cell in enumerate(display_header):
        if _norm(cell) in {"high courts", "high court"}:
            hc_col = idx
            break

    row_maps = []
    for rec in records:
        src_row_idx = rec.get("source_row") or 0
        src_row = rows[src_row_idx - 1] if 0 < src_row_idx <= len(rows) else ()
        row_maps.append({
            "source_row": src_row_idx,
            "source": {
                "High Courts": src_row[hc_col] if hc_col < len(src_row) else None,
                "Component (source col)": rec["component"],
                "Amount (₹)": rec.get("source_rupees"),
            },
            "target": {
                "High Court": rec["high_court"],
                "Component": rec["component"],
                "District": "",
                "Fund Released": rec.get("fund_released"),
                "Fund Utilized": rec.get("fund_utilized"),
                "Remarks": rec.get("remarks") or "",
            },
            "status": "ok",
            "fund_field": fund_key,
        })

    for iss in issues:
        row_maps.append({
            "source_row": iss.get("row"),
            "source": {"error": iss.get("error")},
            "target": {},
            "status": "error",
            "error": iss.get("error"),
        })

    bulk_rows = fin_records_to_bulk_rows(records)
    bulk_bytes = write_bulk_xlsx(FIN_HEADERS, bulk_rows[1:], "Financial_Bulk")

    return {
        "tracker": "financial",
        "format_detected": f"wide_doj_financial_{resolved_mode}",
        "sheet": sheet_name,
        "mode": resolved_mode,
        "column_mappings": column_mappings,
        "row_mappings": row_maps[:MAX_ROW_MAPPINGS],
        "row_mappings_total": len(row_maps),
        "stats": result.get("stats") or {},
        "issues": issues,
        "bulk_bytes": bulk_bytes,
        "bulk_filename": f"financial_{resolved_mode}_converted_bulk.xlsx",
        "desired_headers": list(FIN_HEADERS),
        "alias_hints": {
            "component_aliases": len(COMPONENT_ALIASES),
            "high_court_aliases": len(HC_ALIASES),
        },
    }


def convert_outcome(raw: bytes, *, sheet: Optional[str] = None) -> dict[str, Any]:
    sheets = load_workbook_sheets(raw)
    sheet_name, rows = pick_sheet(sheets, sheet)
    if not rows:
        raise ValueError("Empty sheet")

    header0 = headers_lower(rows[0])
    display_header = header_cells(rows[0])

    if is_long_outcome(header0) and not is_phase4_outcome(header0):
        mappings = [{"source": h, "target": h, "transform": "identity"} for h in OUTCOME_BULK_HEADERS]
        row_maps = []
        for i, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            by = {_norm(display_header[j]): row[j] if j < len(row) else None for j in range(len(display_header))}
            target = {h: by.get(_norm(h)) for h in OUTCOME_BULK_HEADERS}
            if target.get("KPI ID") is None:
                target["KPI ID"] = by.get("kpid")
            row_maps.append({
                "source_row": i,
                "source": _source_snapshot(row, display_header),
                "target": target,
                "status": "ok",
            })
        return {
            "tracker": "outcome",
            "format_detected": "long_bulk",
            "sheet": sheet_name,
            "column_mappings": mappings,
            "row_mappings": row_maps[:MAX_ROW_MAPPINGS],
            "row_mappings_total": len(row_maps),
            "stats": {"records": len(row_maps)},
            "issues": [],
            "bulk_bytes": raw,
            "bulk_filename": "outcome_bulk.xlsx",
            "desired_headers": list(OUTCOME_BULK_HEADERS),
        }

    parsed = parse_outcome_excel_rows(rows)

    column_mappings = [
        {"source": "High Court", "target": "High Court", "transform": "forward-fill blank cells"},
        {
            "source": "Components / Sub-Component",
            "target": "Subject / Component / Sub-Component",
            "transform": "OUTCOME_SUBJECT_MAP",
        },
        {"source": "KPID / KPI ID", "target": "KPI ID", "transform": "normalize"},
        {"source": "Granularity", "target": "Granularity", "transform": "identity (default District)"},
        {"source": "Value / OUTCOME* column", "target": "Value", "transform": "numeric"},
    ]

    row_maps = []
    data_rows: list[list[Any]] = []
    for i, rec in enumerate(parsed, start=2):
        subject = map_outcome_subject(rec.get("subject"))
        target = {
            "High Court": rec.get("high_court"),
            "Component": rec.get("component") or "",
            "Sub-Component": rec.get("sub_component") or "",
            "Subject": subject,
            "KPI ID": rec.get("kpi_id"),
            "Granularity": rec.get("granularity") or "District",
            "District": "",
            "Value": rec.get("value"),
            "Baseline": "",
            "Remarks": "ETL: Outcome convert",
        }
        row_maps.append({
            "source_row": i,
            "source": {
                "High Court": rec.get("high_court"),
                "Subject (mapped)": subject,
                "KPI ID": rec.get("kpi_id"),
                "KPI": rec.get("kpi"),
                "Value": rec.get("value"),
            },
            "target": target,
            "status": "ok",
        })
        data_rows.append([
            target["High Court"],
            target["Component"],
            target["Sub-Component"],
            target["Subject"],
            target["KPI ID"],
            target["Granularity"],
            target["District"],
            target["Value"] if target["Value"] is not None else "",
            target["Baseline"],
            target["Remarks"],
        ])

    bulk_bytes = write_bulk_xlsx(OUTCOME_BULK_HEADERS, data_rows, "Outcome_Bulk")

    return {
        "tracker": "outcome",
        "format_detected": "phase4_outcome",
        "sheet": sheet_name,
        "column_mappings": column_mappings,
        "row_mappings": row_maps[:MAX_ROW_MAPPINGS],
        "row_mappings_total": len(row_maps),
        "stats": {"records": len(parsed)},
        "issues": [],
        "bulk_bytes": bulk_bytes,
        "bulk_filename": "outcome_converted_bulk.xlsx",
        "desired_headers": list(OUTCOME_BULK_HEADERS),
    }


def convert_tracker(
    tracker: str,
    raw: bytes,
    *,
    mode: str = "auto",
    sheet: Optional[str] = None,
) -> dict[str, Any]:
    tracker = (tracker or "").strip().lower()
    if tracker == "physical":
        return convert_physical(raw, sheet=sheet)
    if tracker == "financial":
        return convert_financial(raw, mode=mode, sheet=sheet)
    if tracker == "outcome":
        return convert_outcome(raw, sheet=sheet)
    raise ValueError("tracker must be physical, financial, or outcome")


def public_convert_payload(result: dict[str, Any]) -> dict[str, Any]:
    """Strip binary fields for JSON response."""
    return {k: v for k, v in result.items() if k not in ("bulk_bytes",)}
